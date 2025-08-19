import openpyxl

def read_employees(file_path: str, sheet_name: str = 'employees.xlsx'):
    """Return list of tuples: (first, middle, last, emp_id) from headered Excel."""
    wb = openpyxl.load_workbook(file_path)
    sh = wb[sheet_name]
    # Read header row to map columns
    headers = [c.value for c in next(sh.iter_rows(min_row=1, max_row=1))]
    col_index = {h.strip().lower(): i for i, h in enumerate(headers) if h}
    required = ['first_name','middle_name','last_name','employee_id']
    for r in required:
        if r not in col_index:
            raise ValueError(f"Missing required column '{r}' in Excel. Present: {list(col_index)}")
    data = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        first = row[col_index['first_name']]
        middle = row[col_index['middle_name']]
        last = row[col_index['last_name']]
        emp_id = str(row[col_index['employee_id']]).split('.')[0] if row[col_index['employee_id']] is not None else ''
        data.append((str(first).strip(), str(middle).strip(), str(last).strip(), emp_id))
    return data
